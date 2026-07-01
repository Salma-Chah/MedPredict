# -*- coding: utf-8 -*-
"""
src/export_clean_data.py
------------------------
Charge le dataset brut, applique tout le nettoyage et sauvegarde
le resultat dans data/data_nettoyee.xlsx avec une mise en forme Excel soignee.

Transformations appliquees :
  1. Suppression de patient_id
  2. Decomposition blood_pressure -> bp_sys + bp_dia
  3. Encodage binaire Yes/No -> 1/0 (diabetes, hypertension, readmitted_30_days)
  4. One-hot encoding gender + discharge_destination

Mise en forme Excel :
  - Feuille principale  "Donnees nettoyees"   : donnees + en-tetes couleurs par groupe
  - Feuille secondaire  "Statistiques"        : resume avant/apres + stats descriptives

Usage :
    python -m src.export_clean_data
    python src/export_clean_data.py
"""

import os
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Chemins
# ---------------------------------------------------------------------------
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(PROJECT_ROOT)

DATA_IN  = os.path.join("data", "hospital_readmissions_30k.csv")
DATA_OUT = os.path.join("data", "data_nettoyee.xlsx")
TARGET   = "readmitted_30_days"

# ---------------------------------------------------------------------------
# Palette de couleurs Excel par groupe de colonnes
#   Numeriques originales  : bleu marine   #0D47A1
#   Binaires encodees      : bleu petrole  #006064
#   One-hot gender         : bleu ardoise  #283593
#   One-hot destination    : bleu acier    #01579B
#   Cible                  : bordeaux      #880E4F  (ressort clairement)
# ---------------------------------------------------------------------------
COL_GROUPS = {
    "num"    : {"cols": ["age", "bp_sys", "bp_dia", "cholesterol", "bmi",
                          "medication_count", "length_of_stay"],
                "fill": "0D47A1"},
    "binary" : {"cols": ["diabetes", "hypertension"],
                "fill": "006064"},
    "gender" : {"cols": ["gender_Female", "gender_Male", "gender_Other"],
                "fill": "283593"},
    "dest"   : {"cols": ["dest_Home", "dest_Nursing_Facility", "dest_Rehab"],
                "fill": "01579B"},
    "target" : {"cols": [TARGET],
                "fill": "880E4F"},
}

# Couleurs lignes alternees  (bleu tres pale / blanc)
FILL_EVEN = PatternFill(fill_type="solid", fgColor="E8F4FD")
FILL_ODD  = PatternFill(fill_type="solid", fgColor="FFFFFF")

# ---------------------------------------------------------------------------
# 1. Chargement
# ---------------------------------------------------------------------------

def load_raw(path: str = DATA_IN) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";")
    print(f"[load]    {df.shape[0]:,} lignes x {df.shape[1]} colonnes")
    print(f"          colonnes brutes : {df.columns.tolist()}")
    return df


# ---------------------------------------------------------------------------
# 2. Nettoyage
# ---------------------------------------------------------------------------

def clean(df: pd.DataFrame) -> pd.DataFrame:
    """Applique toutes les transformations et retourne le DataFrame nettoye."""

    df = df.copy()

    # -- Etape 1 : suppression patient_id ---------------------------------
    df.drop(columns=["patient_id"], inplace=True)
    print("[clean]   patient_id supprime")

    # -- Etape 2 : decomposition blood_pressure ---------------------------
    bp = df["blood_pressure"].str.split("/", expand=True).astype(float)
    pos = df.columns.get_loc("blood_pressure")
    df.insert(pos,     "bp_sys", bp[0])
    df.insert(pos + 1, "bp_dia", bp[1])
    df.drop(columns=["blood_pressure"], inplace=True)
    print(f"[clean]   blood_pressure -> bp_sys (moy={df['bp_sys'].mean():.1f})"
          f" + bp_dia (moy={df['bp_dia'].mean():.1f})")

    # -- Etape 3 : encodage binaire Yes/No -> 1/0 -------------------------
    binary_cols = ["diabetes", "hypertension", TARGET]
    for col in binary_cols:
        df[col] = df[col].map({"Yes": 1, "No": 0})
    print(f"[clean]   Yes/No -> 1/0 sur {binary_cols}")

    # -- Etape 4 : one-hot encoding ---------------------------------------
    df = pd.get_dummies(df, columns=["gender"],
                        prefix="gender", drop_first=False, dtype=int)
    df = pd.get_dummies(df, columns=["discharge_destination"],
                        prefix="dest",   drop_first=False, dtype=int)

    gender_cols = [c for c in df.columns if c.startswith("gender_")]
    dest_cols   = [c for c in df.columns if c.startswith("dest_")]
    print(f"[clean]   one-hot gender    -> {gender_cols}")
    print(f"[clean]   one-hot discharge -> {dest_cols}")

    # -- Placer la cible en derniere colonne ------------------------------
    cols = [c for c in df.columns if c != TARGET] + [TARGET]
    df   = df[cols]

    print(f"[clean]   {df.shape[0]:,} lignes x {df.shape[1]} colonnes apres nettoyage")
    print(f"          colonnes finales : {df.columns.tolist()}")
    return df


# ---------------------------------------------------------------------------
# 3. Calcul de la largeur optimale de chaque colonne
# ---------------------------------------------------------------------------

def _optimal_width(col_name: str, series: pd.Series, max_width: int = 28) -> float:
    """
    Calcule la largeur Excel optimale pour une colonne :
    max(longueur de l'en-tete, longueur max des valeurs) + marge.
    Borne entre 8 et max_width.
    """
    header_len = len(col_name)
    # Echantillon de 500 valeurs pour ne pas ralentir sur 30k lignes
    sample = series.dropna().sample(min(500, len(series)), random_state=0)
    data_len = sample.astype(str).str.len().max() if len(sample) > 0 else 0
    width = max(header_len, data_len) + 3
    return float(max(8, min(width, max_width)))


# ---------------------------------------------------------------------------
# 4. Construction des styles de cellule
# ---------------------------------------------------------------------------

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF", name="Calibri", size=11)


def _data_font() -> Font:
    return Font(name="Calibri", size=10)


def _center_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _thin_border(color: str = "CCCCCC") -> Border:
    side = Side(border_style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_border() -> Border:
    white = Side(border_style="thin", color="FFFFFF")
    bottom = Side(border_style="medium", color="FFFFFF")
    return Border(left=white, right=white, top=white, bottom=bottom)


# ---------------------------------------------------------------------------
# 5. Ecriture de la feuille principale "Donnees nettoyees"
# ---------------------------------------------------------------------------

def write_main_sheet(ws, df: pd.DataFrame) -> None:
    """
    Ecrit les donnees sur la feuille ws avec :
      - En-tetes colories par groupe (voir COL_GROUPS)
      - Lignes alternees bleu pale / blanc
      - Bordures fines
      - Largeurs de colonnes automatiques
      - Ligne d'en-tete fixee (freeze)
    """
    # -- Construire un mapping colonne -> couleur d'en-tete
    col_color = {}
    for group in COL_GROUPS.values():
        for col in group["cols"]:
            if col in df.columns:
                col_color[col] = group["fill"]

    header_row  = df.columns.tolist()
    data_font   = _data_font()
    border      = _thin_border()

    # ---- En-tetes --------------------------------------------------------
    for col_idx, col_name in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = _header_fill(col_color.get(col_name, "1565C0"))
        cell.font      = _header_font()
        cell.alignment = _center_align(wrap=True)
        cell.border    = _header_border()

    ws.row_dimensions[1].height = 32

    # ---- Donnees ---------------------------------------------------------
    print("[format]  Ecriture des donnees (peut prendre quelques secondes)...")
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        fill = FILL_EVEN if row_idx % 2 == 0 else FILL_ODD
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = data_font
            cell.alignment = _center_align()
            cell.border    = border

    # ---- Largeurs de colonnes --------------------------------------------
    for col_idx, col_name in enumerate(header_row, start=1):
        col_letter = get_column_letter(col_idx)
        width = _optimal_width(col_name, df[col_name])
        ws.column_dimensions[col_letter].width = width

    # ---- Figer la premiere ligne ----------------------------------------
    ws.freeze_panes = "A2"

    # ---- Filtre automatique sur les en-tetes ----------------------------
    ws.auto_filter.ref = ws.dimensions

    print(f"[format]  Feuille principale : {df.shape[0]:,} lignes x {df.shape[1]} colonnes")


# ---------------------------------------------------------------------------
# 6. Ecriture de la feuille "Statistiques"
# ---------------------------------------------------------------------------

def write_stats_sheet(ws, df_raw: pd.DataFrame, df_clean: pd.DataFrame) -> None:
    """
    Ecrit un resume comparatif avant/apres nettoyage et les stats descriptives.
    """
    BLUE    = _header_fill("1565C0")
    GREEN   = _header_fill("1B5E20")
    ORANGE  = _header_fill("E65100")
    h_font  = _header_font()
    d_font  = _data_font()
    border  = _thin_border()
    center  = _center_align()
    FILL_G  = PatternFill(fill_type="solid", fgColor="F1F8E9")   # vert pale
    FILL_B  = PatternFill(fill_type="solid", fgColor="E3F2FD")   # bleu pale

    row = 1

    # ---- Bloc 1 : comparaison avant / apres -----------------------------
    def write_section_title(ws, row, title, fill, ncols=4):
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = fill
        cell.font = h_font
        cell.alignment = center
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
        ws.row_dimensions[row].height = 22
        return row + 1

    row = write_section_title(ws, row, "COMPARAISON AVANT / APRES NETTOYAGE", BLUE, ncols=4)

    # En-tetes
    headers = ["", "Avant nettoyage", "Apres nettoyage", "Delta"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _header_fill("1976D2")
        cell.font = h_font
        cell.alignment = center
    row += 1

    comparisons = [
        ("Lignes",    df_raw.shape[0],   df_clean.shape[0],
         df_clean.shape[0] - df_raw.shape[0]),
        ("Colonnes",  df_raw.shape[1],   df_clean.shape[1],
         df_clean.shape[1] - df_raw.shape[1]),
        ("Valeurs nulles", int(df_raw.isnull().sum().sum()),
         int(df_clean.isnull().sum().sum()),
         int(df_clean.isnull().sum().sum()) - int(df_raw.isnull().sum().sum())),
        ("Doublons",  int(df_raw.duplicated().sum()),
         int(df_clean.duplicated().sum()),
         int(df_clean.duplicated().sum()) - int(df_raw.duplicated().sum())),
    ]
    for i, (label, before, after, delta) in enumerate(comparisons):
        fill = FILL_B if i % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFF")
        delta_str = f"+{delta}" if delta > 0 else str(delta)
        for c, val in enumerate([label, before, after, delta_str], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill
            cell.font      = d_font
            cell.alignment = center
            cell.border    = border
        row += 1

    row += 1  # ligne vide

    # ---- Bloc 2 : types des colonnes finales ----------------------------
    row = write_section_title(ws, row, "TYPES DES COLONNES APRES NETTOYAGE", GREEN, ncols=4)

    for c, h in enumerate(["Colonne", "Type", "Nb unique", "Groupe"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _header_fill("2E7D32")
        cell.font = h_font
        cell.alignment = center
    row += 1

    # Mapping colonne -> groupe
    col_to_group = {}
    group_labels = {
        "num":    "Numerique",
        "binary": "Binaire encode (0/1)",
        "gender": "One-hot gender",
        "dest":   "One-hot destination",
        "target": "Variable cible",
    }
    for gname, gdata in COL_GROUPS.items():
        for col in gdata["cols"]:
            col_to_group[col] = group_labels[gname]

    for i, col in enumerate(df_clean.columns):
        fill = FILL_G if i % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFF")
        vals = [col,
                str(df_clean[col].dtype),
                int(df_clean[col].nunique()),
                col_to_group.get(col, "")]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill
            cell.font      = d_font
            cell.alignment = center
            cell.border    = border
        row += 1

    row += 1  # ligne vide

    # ---- Bloc 3 : statistiques descriptives ----------------------------
    row = write_section_title(ws, row, "STATISTIQUES DESCRIPTIVES — COLONNES NUMERIQUES",
                               ORANGE, ncols=8)

    num_cols = COL_GROUPS["num"]["cols"]
    stat_headers = ["Colonne", "Min", "Max", "Moyenne", "Mediane", "Ecart-type", "Q1", "Q3"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _header_fill("BF360C")
        cell.font = h_font
        cell.alignment = center
    row += 1

    for i, col in enumerate(num_cols):
        if col not in df_clean.columns:
            continue
        s = df_clean[col]
        fill = PatternFill(fill_type="solid", fgColor="FBE9E7") \
               if i % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFF")
        vals = [col,
                round(float(s.min()),  2),
                round(float(s.max()),  2),
                round(float(s.mean()), 3),
                round(float(s.median()), 3),
                round(float(s.std()),  3),
                round(float(s.quantile(0.25)), 3),
                round(float(s.quantile(0.75)), 3)]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill
            cell.font      = d_font
            cell.alignment = center
            cell.border    = border
        row += 1

    row += 1

    # ---- Bloc 4 : desequilibre cible ------------------------------------
    row = write_section_title(ws, row, "REPARTITION DE LA VARIABLE CIBLE (readmitted_30_days)",
                               _header_fill("4A148C"), ncols=4)

    for c, h in enumerate(["Classe", "Libelle", "Count", "Pourcentage (%)"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _header_fill("6A1B9A")
        cell.font = h_font
        cell.alignment = center
    row += 1

    counts = df_clean[TARGET].value_counts().sort_index()
    labels = {0: "Non readmis (No)", 1: "Readmis (Yes)"}
    colors = ["E8F5E9", "FCE4EC"]
    for i, (cls, cnt) in enumerate(counts.items()):
        pct = cnt / len(df_clean) * 100
        fill = PatternFill(fill_type="solid", fgColor=colors[i])
        for c, val in enumerate([int(cls), labels[cls], int(cnt), round(pct, 2)], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill
            cell.font      = d_font
            cell.alignment = center
            cell.border    = border
        row += 1

    # ---- Largeurs colonnes de la feuille stats -------------------------
    widths = [28, 18, 18, 24, 18, 18, 18, 18]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    print("[format]  Feuille Statistiques : OK")


# ---------------------------------------------------------------------------
# 7. Sauvegarde Excel
# ---------------------------------------------------------------------------

def export_to_excel(
    df_raw: pd.DataFrame,
    df_clean: pd.DataFrame,
    path: str = DATA_OUT,
) -> None:
    """Ecrit les deux feuilles dans le fichier Excel et applique la mise en forme."""

    print(f"\n[export]  Ecriture vers {path} ...")

    # Ecriture initiale avec pandas (creer le fichier + la feuille principale)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_clean.to_excel(
            writer,
            sheet_name="Donnees nettoyees",
            index=False,
        )
        # Feuille stats vide — sera remplie apres
        pd.DataFrame().to_excel(writer, sheet_name="Statistiques", index=False)

    # Reouverture pour appliquer la mise en forme (openpyxl)
    wb = load_workbook(path)

    # -- Feuille principale
    ws_main = wb["Donnees nettoyees"]
    write_main_sheet(ws_main, df_clean)

    # -- Feuille statistiques
    ws_stats = wb["Statistiques"]
    write_stats_sheet(ws_stats, df_raw, df_clean)

    # -- Proprietes du classeur
    wb.properties.title   = "Hospital Readmission — Donnees Nettoyees"
    wb.properties.subject = "Preprocessing PFA"
    wb.properties.creator = "src/export_clean_data.py"

    wb.save(path)
    size_kb = os.path.getsize(path) / 1024
    print(f"[export]  Sauvegarde OK -> {path}  ({size_kb:.0f} Ko)")


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------

def export_pipeline(
    data_in:  str = DATA_IN,
    data_out: str = DATA_OUT,
) -> pd.DataFrame:

    print("=" * 60)
    print("EXPORT DONNEES NETTOYEES -> EXCEL")
    print("=" * 60)

    df_raw   = load_raw(data_in)
    df_clean = clean(df_raw)

    export_to_excel(df_raw, df_clean, data_out)

    print("\n" + "=" * 60)
    print("RESUME")
    print("=" * 60)
    print(f"  Fichier source  : {data_in}")
    print(f"  Fichier sortie  : {data_out}")
    print(f"  Feuilles        : 'Donnees nettoyees' + 'Statistiques'")
    print(f"  Lignes          : {df_clean.shape[0]:,}")
    print(f"  Colonnes        : {df_clean.shape[1]}  ({', '.join(df_clean.columns.tolist())})")
    print("=" * 60)

    return df_clean


# ---------------------------------------------------------------------------
# Point d'entree
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    export_pipeline()
